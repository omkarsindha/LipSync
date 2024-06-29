import openpyxl
from openpyxl import utils
from openpyxl.styles import Alignment

LINK_TYPES = ['SD', 'HD', '3GA']
LINK_LINE_SUPPORTED = {
    'SD': ['525i', '625i'],
    'HD': ['720p', '1080i', '1080sF', '1080p'],
    '3GA': ['720p', '1080i', '1080sF', '1080p']
}
FRAME_RATES_SUPPORTED = {  # Dictionary matching format with list of its valid frame-rates
    '525i': ['59.94'],
    '625i': ['50'],
    '720p': ['23.98', '24', '25', '29.97', '30', '50', '59.94', '60'],
    '1080i': ['50', '59.94', '60'],
    '1080sF': ['23.98', '24', '25', '29.97', '30'],
    '1080p': ['23.98', '24', '25', '29.97', '30']
}
PHABRIX_VALUES_LINK = {  # Dictionary to match phabrix values to the actual option
    'SD': 0,
    'HD': 1,
    '3GA': 3
}
PHABRIX_VALUES_LINE = {
    '525i': 0,
    '625i': 1,
    '720p': 2,
    '1080i': 4,
    '1080sF': 5,
    '1080p': 6
}
PHABRIX_VALUES_FRAME_RATE = {
    '23.98': 0,
    '24': 1,
    '25': 2,
    '29.97': 3,
    '30': 4,
    '50': 5,
    '59.94': 6,
    '60': 7
}
# Different IPG options
IPGS = ['570 A9', '570 X-19', 'evIPG-12G', 'evIPG-3G']


class Config:

    def __init__(self):
        self.PHABRIX_VALUE = []      # Stores the actual value of the FORMAT to send msg to phabrix eg, 1,3,5
        self.FORMATS = []            # Used for getting the string version of phabrix values HD 1080i 59.94
        self.OUTPUT_AV_SYNC = []     # Used to store different values user wants to set for IPG's Output AV Sync mode
        self.VERTICAL_OFFSET = []    # " for IPG's vertical offset
        self.AES67 = []              # " for AES67 modes
        self.OUTS = []               # " outputs that user wants to test
        self.DELAY = 0               # delay between each test
        self.test_result = []        # Stores the result (used to save result as excel)
        self.EXPECTED = {}           # Keeps track of the expected values from the file
        self.load_config()           # Loads values from testconfig.txt
        self.load_expected()         # Loads values from expected.txt

    def load_expected(self):
        self.EXPECTED = {}
        with open("Config/expected.txt", "r") as file:
            for line in file:
                if "//" in line:   # It is a comment
                    continue
                if "END" in line:  # End of the file
                    break
                row = line.strip().split(" ")
                key = f"{row[0]}{row[1]}{row[2]}{row[3]}{row[4]}"       # Makes a key out of the expected value's config
                value1 = float(row[5])
                value2 = float(row[6])
                self.EXPECTED[key] = [value1, value2]                   # Stores the right and left value as a list

    def load_config(self):
        self.PHABRIX_VALUE: list = []
        self.OUTS: list = []
        self.OUTPUT_AV_SYNC: list = []
        self.VERTICAL_OFFSET: list = []
        self.AES67: list = []
        self.FORMATS: list = []
        self.DELAY: int = 0
        record: bool = False
        with open("Config/testconfig.txt", "r") as file:
            for line in file:
                # Checking if the line is a comment or not
                if "//" in line:
                    continue

                if "DELAY" in line:
                    self.DELAY = int(next(file).strip())

                if "IPGOUTS" in line:
                    self.OUTS = [int(x) for x in next(file).strip().split(' ')]  # Converts them to integer

                if "OUTPUT_AV_SYNC" in line:
                    self.OUTPUT_AV_SYNC = [int(x) for x in next(file).strip().split(' ')]  # Converts them to integer

                if "VERTICAL_OFFSET" in line:
                    self.VERTICAL_OFFSET = [int(x) for x in next(file).strip().split(' ')]  # Converts them to integer

                if "AES" in line:
                    self.AES67 = [int(x) for x in next(file).strip().split(' ')]  # Converts them to integer

                if "END_RECORD" in line:
                    record = False
                if record:
                    # Checking if user made a mistake in the configuration file
                    FORMAT = line.strip().split(' ')
                    LINK = PHABRIX_VALUES_LINK.get(FORMAT[0], 'INVALID')        # If found it is valid otherwise Invalid
                    LINE = PHABRIX_VALUES_LINE.get(FORMAT[1], 'INVALID')        # "
                    RATE = PHABRIX_VALUES_FRAME_RATE.get(FORMAT[2], 'INVALID')  # "
                    if LINK != "INVALID" and LINE != "INVALID" and RATE != "INVALID":  # If all 3 valid
                        # Checking if format exists or not
                        if FORMAT[1] in LINK_LINE_SUPPORTED[FORMAT[0]] and FORMAT[2] in FRAME_RATES_SUPPORTED[FORMAT[1]]:
                            # Line and rate are valid for their Link and Line respectively
                            self.PHABRIX_VALUE.append([LINK, LINE, RATE])  # Appending them as is as it is valid
                            self.FORMATS.append(FORMAT)
                        else:
                            self.PHABRIX_VALUE.append(['INVALID', 'INVALID', 'INVALID'])  # Make all 3 invalid
                            self.FORMATS.append(["Not a valid format", "", "(Check testcongif file)"])
                    else:
                        self.FORMATS.append(["INVALID", "", "(Check testconfig file)"])

                if "STANDARDS" in line:
                    record = True

    def save_config(self, filename):
        HEADINGS = ["#", "IPG Out Tested", "Format on Phabrix", "OutputAv", "Vertical Offset", "AES", "Delay Right",
                    "Delay Left", "Min", "Max", "Result"]
        workbook = openpyxl.Workbook()

        get_col_let = utils.get_column_letter
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        worksheet = workbook.create_sheet(title="IPG Lip Sync test results")

        worksheet.column_dimensions[get_col_let(1)].width = 3
        worksheet.column_dimensions[get_col_let(2)].width = 16
        worksheet.column_dimensions[get_col_let(3)].width = 18
        worksheet.column_dimensions[get_col_let(4)].width = 12
        worksheet.column_dimensions[get_col_let(5)].width = 13
        worksheet.column_dimensions[get_col_let(6)].width = 5.5
        worksheet.column_dimensions[get_col_let(7)].width = 12
        worksheet.column_dimensions[get_col_let(8)].width = 12
        worksheet.column_dimensions[get_col_let(9)].width = 12
        worksheet.column_dimensions[get_col_let(10)].width = 12

        worksheet.append(HEADINGS)

        for result in self.test_result:
            worksheet.append(result)

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left')

        workbook.save(filename)


if __name__ == "__main__":
    config_instance = Config()
